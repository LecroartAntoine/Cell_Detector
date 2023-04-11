from PyQt5 import QtCore, QtGui, QtWidgets
import os, cv2, psutil, ImageViewer, sys, re, utils
import numpy as np

class DetectThread(QtCore.QThread):
    finished = QtCore.pyqtSignal(object)

    def __init__(self, images, taux_conf, box_width, show_conf, show_name, single = None, parent=None):
        super().__init__(parent)
        self.images = images
        self.taux_conf = taux_conf
        self.box_width = box_width
        self.show_conf = show_conf
        self.show_name = show_name
        self.single = single

    def run(self):
        if self.single:
            self.images[self.single]['pred'] = utils.yolo_detection(self.images[self.single]['image'], self.taux_conf / 100)
            self.images[self.single]['image_pred'] = utils.plot_bboxes(self.images[self.single]['image'], self.images[self.single]['pred'].boxes.boxes, self.box_width, self.show_conf, self.show_name)

        else:
            for key in self.images:
                self.images[key]['pred'] = utils.yolo_detection(self.images[key]['image'], self.taux_conf / 100)
                self.images[key]['nb_cell'] = len(self.images[key]['pred'].boxes.boxes)
                self.images[key]['image_pred'] = utils.plot_bboxes(self.images[key]['image'], self.images[key]['pred'].boxes.boxes, self.box_width, self.show_conf, self.show_name)
        
        self.finished.emit(self.images)

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1200, 800)

        # self.images = {}

        # Image Viewer
        self.detect_viewer = ImageViewer.PhotoViewer(MainWindow)
        self.analyse_viewer = ImageViewer.PhotoViewer(MainWindow)

        # Main window
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.mainLayout = QtWidgets.QFormLayout(self.centralwidget)
        self.mainLayout.setFieldGrowthPolicy(QtWidgets.QFormLayout.AllNonFixedFieldsGrow)
        self.mainLayout.setObjectName("mainLayout")

#################################  MENU  ##################################################
        self.menu = QtWidgets.QVBoxLayout()
        self.menu.setObjectName("menu")

        self.logo = QtWidgets.QLabel(self.centralwidget)
        self.logo.setObjectName('Logo')
        self.logo.setPixmap(QtGui.QPixmap("Assets:Logo.png"))
        self.menu.addWidget(self.logo)


        self.menu_button = QtWidgets.QVBoxLayout()
        self.menu_button.setContentsMargins(15, -1, 15, -1)
        self.menu_button.setObjectName("menu_button")

        self.open = QtWidgets.QPushButton(self.centralwidget)
        self.open.setObjectName("open")
        self.open.clicked.connect(self.openfile)
        self.open.setIcon(QtGui.QIcon('Assets:open.png'))
        self.open.setIconSize(QtCore.QSize(30, 30))
        self.menu_button.addWidget(self.open)

        self.select_button = QtWidgets.QPushButton(self.centralwidget)
        self.select_button.setObjectName("select_button")
        self.select_button.clicked.connect(self.change_page_1)
        self.select_button.setEnabled(False)
        self.select_button.setIcon(QtGui.QIcon('Assets:select.png'))
        self.select_button.setIconSize(QtCore.QSize(30, 30))
        self.menu_button.addWidget(self.select_button)

        self.detection_button = QtWidgets.QPushButton(self.centralwidget)
        self.detection_button.setObjectName("detection_button")
        self.detection_button.clicked.connect(self.change_page_2)
        self.detection_button.setEnabled(False)
        self.detection_button.setIcon(QtGui.QIcon('Assets:detect.png'))
        self.detection_button.setIconSize(QtCore.QSize(30, 30))
        self.menu_button.addWidget(self.detection_button)

        self.analyse_button = QtWidgets.QPushButton(self.centralwidget)
        self.analyse_button.setObjectName("analyse_button")
        self.analyse_button.clicked.connect(self.change_page_3)
        self.analyse_button.setEnabled(False)
        self.analyse_button.setIcon(QtGui.QIcon('Assets:analyse.png'))
        self.analyse_button.setIconSize(QtCore.QSize(30, 30))
        self.menu_button.addWidget(self.analyse_button)

        self.calculs_button = QtWidgets.QPushButton(self.centralwidget)
        self.calculs_button.setObjectName("calculs_button")
        self.calculs_button.clicked.connect(self.change_page_4)
        self.calculs_button.setEnabled(False)
        self.calculs_button.setIcon(QtGui.QIcon('Assets:calcul.png'))
        self.calculs_button.setIconSize(QtCore.QSize(30, 30))
        self.menu_button.addWidget(self.calculs_button)

        self.files = QtWidgets.QListWidget(self.centralwidget)
        self.files.setObjectName("files")
        self.files.currentRowChanged.connect(self.change_image)
        self.menu_button.addWidget(self.files)

        self.menu.addLayout(self.menu_button)

        self.cpu = QtWidgets.QHBoxLayout()
        self.cpu.setObjectName("cpu")
        self.cpu_label = QtWidgets.QLabel(self.centralwidget)
        self.cpu_label.setObjectName("cpu_label")
        self.cpu.addWidget(self.cpu_label)
        self.cpu_bar = QtWidgets.QProgressBar(self.centralwidget)
        self.cpu_bar.setProperty("value", 1)
        self.cpu_bar.setObjectName("cpu_bar")
        self.cpu.addWidget(self.cpu_bar)
        self.menu_button.addLayout(self.cpu)
        self.ram = QtWidgets.QHBoxLayout()
        self.ram.setObjectName("ram")
        self.ram_label = QtWidgets.QLabel(self.centralwidget)
        self.ram_label.setObjectName("ram_label")
        self.ram.addWidget(self.ram_label)
        self.ram_bar = QtWidgets.QProgressBar(self.centralwidget)
        self.ram_bar.setProperty("value", 1)
        self.ram_bar.setObjectName("ram_bar")
        self.ram.addWidget(self.ram_bar)
        self.menu_button.addLayout(self.ram)

        self.mainLayout.setLayout(1, QtWidgets.QFormLayout.LabelRole, self.menu)


#####################################  Browser  ########################################################
        self.browser = QtWidgets.QStackedWidget(self.centralwidget)
        self.browser.setObjectName("browser")
        self.browser.setCurrentIndex(0)

##################################### Welcome page ########################################################

        self.welcome = QtWidgets.QWidget()
        self.welcome.setObjectName("welcome")
        self.welcome_layout = QtWidgets.QVBoxLayout(self.welcome)
        self.welcome_layout.setObjectName("welcome_layout")

        self.welcome_message_1 = QtWidgets.QLabel(self.welcome)
        self.welcome_message_1.setAlignment(QtCore.Qt.AlignCenter)
        self.welcome_message_1.setObjectName("welcome_message_1")

        self.welcome_message_2 = QtWidgets.QLabel(self.welcome)
        self.welcome_message_2.setAlignment(QtCore.Qt.AlignCenter)
        self.welcome_message_2.setObjectName("welcome_message_2")

        self.welcome_layout.addWidget(self.welcome_message_1)
        self.welcome_layout.addWidget(self.welcome_message_2)
        self.browser.addWidget(self.welcome)

##################################### Selection  ########################################################
        self.select = QtWidgets.QWidget()
        self.select.setObjectName("select")
        self.image_select_layout = QtWidgets.QVBoxLayout(self.select)
        self.image_select_layout.setContentsMargins(0, 0, 0, 0)
        self.image_select_layout.setObjectName("image_select_layout")
        self.image_type_layout = QtWidgets.QGridLayout()
        self.image_type_layout.setObjectName("image_type_layout")
        self.type_2s = QtWidgets.QComboBox(self.select)
        self.type_2s.setObjectName("type_2s")
        self.image_type_layout.addWidget(self.type_2s, 3, 1, 1, 1)
        self.type_1c = QtWidgets.QComboBox(self.select)
        self.type_1c.setObjectName("type_1c")
        self.image_type_layout.addWidget(self.type_1c, 6, 0, 1, 1)
        self.type_2c = QtWidgets.QComboBox(self.select)
        self.type_2c.setObjectName("type2c")
        self.image_type_layout.addWidget(self.type_2c, 6, 1, 1, 1)
        self.type_1s = QtWidgets.QComboBox(self.select)
        self.type_1s.setObjectName("type_1s")
        self.image_type_layout.addWidget(self.type_1s, 3, 0, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.image_type_layout.addItem(spacerItem, 4, 0, 1, 1)
        self.type_2c_label = QtWidgets.QLabel(self.select)
        self.type_2c_label.setObjectName("type_2c_label")
        self.image_type_layout.addWidget(self.type_2c_label, 5, 1, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.image_type_layout.addItem(spacerItem, 4, 1, 1, 1)
        self.type_2s_label = QtWidgets.QLabel(self.select)
        self.type_2s_label.setObjectName("type_2s_label")
        self.image_type_layout.addWidget(self.type_2s_label, 2, 1, 1, 1)
        self.type_3c = QtWidgets.QComboBox(self.select)
        self.type_3c.setObjectName("type_3c")
        self.image_type_layout.addWidget(self.type_3c, 6, 2, 1, 1)
        self.type_1c_label = QtWidgets.QLabel(self.select)
        self.type_1c_label.setObjectName("type_1c_label")
        self.image_type_layout.addWidget(self.type_1c_label, 5, 0, 1, 1)
        self.type_1s_label = QtWidgets.QLabel(self.select)
        self.type_1s_label.setObjectName("type_1s_label")
        self.image_type_layout.addWidget(self.type_1s_label, 2, 0, 1, 1)
        self.type_3s_label = QtWidgets.QLabel(self.select)
        self.type_3s_label.setObjectName("type_3s_label")
        self.image_type_layout.addWidget(self.type_3s_label, 2, 2, 1, 1)
        self.type_3s = QtWidgets.QComboBox(self.select)
        self.type_3s.setObjectName("type_3s")
        self.image_type_layout.addWidget(self.type_3s, 3, 2, 1, 1)
        self.type_3c_label = QtWidgets.QLabel(self.select)
        self.type_3c_label.setObjectName("type_3c_label")
        self.image_type_layout.addWidget(self.type_3c_label, 5, 2, 1, 1)
        self.image_select_label = QtWidgets.QLabel(self.select)
        self.image_select_label.setAlignment(QtCore.Qt.AlignCenter)
        self.image_select_label.setObjectName("image_select_label")
        self.image_type_layout.addWidget(self.image_select_label, 0, 1, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.image_type_layout.addItem(spacerItem, 4, 2, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.image_type_layout.addItem(spacerItem, 1, 1, 1, 1)
        self.image_type_layout.setRowMinimumHeight(0, 60)
        self.image_type_layout.setRowMinimumHeight(1, 60)
        self.image_type_layout.setRowMinimumHeight(2, 60)
        self.image_type_layout.setRowMinimumHeight(3, 60)
        self.image_type_layout.setRowMinimumHeight(4, 60)
        self.image_select_layout.addLayout(self.image_type_layout)

        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.image_select_layout.addItem(spacerItem)

        self.cycle_layout = QtWidgets.QHBoxLayout()
        self.cycle_layout.setObjectName("cycle_layout")

        self.cycle_prev_button = QtWidgets.QPushButton(self.select)
        self.cycle_prev_button.setObjectName("cycle_prev_button")
        self.cycle_prev_button.clicked.connect(self.cycle_prev)
        self.cycle_prev_button.setEnabled(False)
        self.cycle_prev_button.setIcon(QtGui.QIcon('Assets:left.png'))
        self.cycle_prev_button.setIconSize(QtCore.QSize(30, 30))
        self.cycle_layout.addWidget(self.cycle_prev_button)

        self.cycle_next_button = QtWidgets.QPushButton(self.select)
        self.cycle_next_button.setObjectName("cycle_next_button")
        self.cycle_next_button.clicked.connect(self.cycle_next)
        self.cycle_next_button.setEnabled(False)
        self.cycle_next_button.setIcon(QtGui.QIcon('Assets:right.png'))
        self.cycle_next_button.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.cycle_next_button.setIconSize(QtCore.QSize(30, 30))
        self.cycle_layout.addWidget(self.cycle_next_button)

        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.cycle_layout.addItem(spacerItem)

        self.image_select_confirm = QtWidgets.QPushButton(self.select)
        self.image_select_confirm.setObjectName("image_select_confirm")
        self.image_select_confirm.setEnabled(False)
        self.image_select_confirm.setIcon(QtGui.QIcon('Assets:confirm.png'))
        self.image_select_confirm.setIconSize(QtCore.QSize(30, 30))
        self.image_select_confirm.clicked.connect(self.set_file_list)
        self.cycle_layout.addWidget(self.image_select_confirm)

        self.image_select_layout.addLayout(self.cycle_layout)

        self.browser.addWidget(self.select)

#####################################  Yolo detection  ########################################################
        self.detect = QtWidgets.QWidget()
        self.detect.setObjectName("detect")
        self.detect_layout = QtWidgets.QVBoxLayout(self.detect)
        self.detect_layout.setObjectName("detect_layout")
        self.detect_tool_layout = QtWidgets.QHBoxLayout()
        self.detect_tool_layout.setObjectName("detect_tool_layout")
        self.detect_tool_layout.setSpacing(20)
        self.detect_model = QtWidgets.QHBoxLayout()
        self.detect_model.setObjectName("detect_model")

        self.show_layout = QtWidgets.QGridLayout()
        self.show_layout.setObjectName("show_layout")
        
        self.show_conf_label = QtWidgets.QLabel(self.detect)
        self.show_conf_label.setObjectName("show_conf_label")
        self.show_layout.addWidget(self.show_conf_label, 0, 0)

        self.show_conf = QtWidgets.QRadioButton(self.detect)
        self.show_conf.setObjectName("show_conf")
        self.show_layout.addWidget(self.show_conf, 0, 1)
        
        self.show_name_label = QtWidgets.QLabel(self.detect)
        self.show_name_label.setObjectName("show_name_label")
        self.show_layout.addWidget(self.show_name_label, 1, 0)

        self.show_name = QtWidgets.QRadioButton(self.detect)
        self.show_name.setObjectName("show_name")
        self.show_layout.addWidget(self.show_name, 1, 1)

        self.detect_tool_layout.addLayout(self.show_layout)

        self.group = QtWidgets.QButtonGroup()
        self.group.addButton(self.show_conf)
        self.group.addButton(self.show_name)       
        self.group.setExclusive(False)
        self.show_name.setChecked(True)
        self.show_conf.setChecked(True)

        self.detect_box = QtWidgets.QHBoxLayout()
        self.detect_box.setObjectName("detect_box")
        self.box_width_label = QtWidgets.QLabel(self.detect)
        self.box_width_label.setObjectName("box_width_label")
        self.detect_box.addWidget(self.box_width_label)
        self.box_width = QtWidgets.QSpinBox(self.detect)
        self.box_width.setProperty("value", 1)
        self.box_width.setObjectName("box_width")
        self.detect_box.addWidget(self.box_width)
        self.detect_tool_layout.addLayout(self.detect_box)

        self.taux_conf_layout  = QtWidgets.QHBoxLayout()
        self.taux_conf_layout.setObjectName("taux_conf_layout")
        self.taux_conf_label = QtWidgets.QLabel(self.detect)
        self.taux_conf_label.setObjectName("taux_conf_label")
        self.taux_conf_layout.addWidget(self.taux_conf_label)
        self.taux_conf = QtWidgets.QSpinBox(self.detect)
        self.taux_conf.setMaximum(100)
        self.taux_conf.setAccelerated(True)
        self.taux_conf.setProperty("value", 5)
        self.taux_conf.setObjectName("taux_conf")
        self.taux_conf_layout.addWidget(self.taux_conf)
        self.detect_tool_layout.addLayout(self.taux_conf_layout)

        self.detect_button_layout = QtWidgets.QVBoxLayout()
        self.detect_button_layout.setObjectName("detect_button_layout")
        self.start_detect_one = QtWidgets.QPushButton(self.detect)
        self.start_detect_one.setObjectName("start_detect_one")
        self.start_detect_one.clicked.connect(lambda : self.get_pred(True))
        self.start_detect_one.setIcon(QtGui.QIcon('Assets:one.png'))
        self.start_detect_one.setIconSize(QtCore.QSize(30, 30))
        self.detect_button_layout.addWidget(self.start_detect_one)

        self.start_detect_many = QtWidgets.QPushButton(self.detect)
        self.start_detect_many.setObjectName("start_detect_many")
        self.start_detect_many.clicked.connect(self.get_pred)
        self.start_detect_many.setIcon(QtGui.QIcon('Assets:many.png'))
        self.start_detect_many.setIconSize(QtCore.QSize(30, 30))
        self.detect_button_layout.addWidget(self.start_detect_many)
        self.detect_tool_layout.addLayout(self.detect_button_layout)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.detect_tool_layout.addItem(spacerItem)
        self.detect_layout.addLayout(self.detect_tool_layout)
        self.detect_layout.addWidget(self.detect_viewer)
        self.browser.addWidget(self.detect)

#####################################  Analyse  ########################################################
        self.analyse = QtWidgets.QWidget()
        self.analyse.setObjectName("analyse")

        self.analyse_layout = QtWidgets.QVBoxLayout(self.analyse)
        self.analyse_layout.setObjectName("analyse_layout")

        self.analyse_tool = QtWidgets.QHBoxLayout()
        self.analyse_tool.setObjectName("analyse_tool")
        self.analyse_tool.setSpacing(20)

        self.show_detection_button = QtWidgets.QPushButton(self.analyse)
        self.show_detection_button.setObjectName("show_detection_button")
        self.show_detection_button.clicked.connect(self.show_detection)
        self.show_detection_button.setIcon(QtGui.QIcon('Assets:grid.png'))
        self.show_detection_button.setIconSize(QtCore.QSize(30, 30))
        self.analyse_tool.addWidget(self.show_detection_button)

        self.show_pred_button = QtWidgets.QPushButton(self.analyse)
        self.show_pred_button.setObjectName("show_pred_button")
        self.show_pred_button.clicked.connect(self.show_pred)
        self.show_pred_button.setEnabled(False)
        self.show_pred_button.setIcon(QtGui.QIcon('Assets:image.png'))
        self.show_pred_button.setIconSize(QtCore.QSize(30, 30))
        self.analyse_tool.addWidget(self.show_pred_button)

        self.analyse_layout.addLayout(self.analyse_tool)
        self.analyse_layout.addWidget(self.analyse_viewer)
        self.browser.addWidget(self.analyse)


        self.mainLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.browser)

#####################################  Calculs  ########################################################

        self.calculs = QtWidgets.QWidget()
        self.calculs.setObjectName("calculs")

        self.calculs_layout = QtWidgets.QVBoxLayout(self.calculs)
        self.calculs_layout.setObjectName("calculs_layout")

        self.calc_buttons_layout = QtWidgets.QHBoxLayout()
        self.calc_buttons_layout.setObjectName("Calc_buttons_layout")


        self.start_calc_button = QtWidgets.QPushButton(self.calculs)
        self.start_calc_button.setObjectName("start_calc_button")
        self.start_calc_button.clicked.connect(self.start_calc)
        self.start_calc_button.setIcon(QtGui.QIcon('Assets:start.png'))
        self.start_calc_button.setIconSize(QtCore.QSize(30, 30))
        self.calc_buttons_layout.addWidget(self.start_calc_button)

        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.calc_buttons_layout.addItem(spacerItem)

        self.img_size_layout = QtWidgets.QHBoxLayout()
        self.img_size_layout.setObjectName("img_size_layout")
        self.img_size_layout.setSpacing(15)

        self.img_size_label = QtWidgets.QLabel(self.calculs)
        self.img_size_label.setObjectName("img_size_label")
        self.calc_buttons_layout.addWidget(self.img_size_label)

        self.img_size_40 = QtWidgets.QRadioButton(self.calculs)
        self.img_size_40.setObjectName("img_size_40")
        self.img_size_layout.addWidget(self.img_size_40)

        self.img_size_100 = QtWidgets.QRadioButton(self.calculs)
        self.img_size_100.setObjectName("img_size_100")
        self.img_size_layout.addWidget(self.img_size_100)

        self.calc_buttons_layout.addLayout(self.img_size_layout)

        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.calc_buttons_layout.addItem(spacerItem)

        self.calc_button_save_xlsx = QtWidgets.QPushButton(self.calculs)
        self.calc_button_save_xlsx.setObjectName("calc_button_save_xlsx")
        self.calc_button_save_xlsx.setEnabled(False)
        self.calc_button_save_xlsx.clicked.connect(self.save_table_xlsx)
        self.calc_button_save_xlsx.setIcon(QtGui.QIcon('Assets:xlsx.png'))
        self.calc_button_save_xlsx.setIconSize(QtCore.QSize(30, 30))
        self.calc_buttons_layout.addWidget(self.calc_button_save_xlsx)

        self.calc_button_save_csv = QtWidgets.QPushButton(self.calculs)
        self.calc_button_save_csv.setObjectName("calc_button_save_csv")
        self.calc_button_save_csv.setEnabled(False)
        self.calc_button_save_csv.clicked.connect(self.save_table_csv)
        self.calc_button_save_csv.setIcon(QtGui.QIcon('Assets:csv.png'))
        self.calc_button_save_csv.setIconSize(QtCore.QSize(30, 30))
        self.calc_buttons_layout.addWidget(self.calc_button_save_csv)
        self.calculs_layout.addLayout(self.calc_buttons_layout)

        self.calculs_table = QtWidgets.QTableWidget(self.calculs)
        self.calculs_table.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        self.calculs_table.setObjectName("calculs_table")
        self.calculs_table.setColumnCount(4)
        self.calculs_table.setRowCount(6)
        QtWidgets.QTableWidgetItem().setTextAlignment(4)
        self.calculs_table.setVerticalHeaderItem(0, QtWidgets.QTableWidgetItem())
        self.calculs_table.setVerticalHeaderItem(1, QtWidgets.QTableWidgetItem())
        self.calculs_table.setVerticalHeaderItem(2, QtWidgets.QTableWidgetItem())
        self.calculs_table.setVerticalHeaderItem(3, QtWidgets.QTableWidgetItem())
        self.calculs_table.setVerticalHeaderItem(4, QtWidgets.QTableWidgetItem())
        self.calculs_table.setVerticalHeaderItem(5, QtWidgets.QTableWidgetItem())
        self.calculs_table.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem())
        self.calculs_table.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem())
        self.calculs_table.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem())
        self.calculs_table.setHorizontalHeaderItem(3, QtWidgets.QTableWidgetItem())
        col_header = self.calculs_table.horizontalHeader()       
        col_header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        col_header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        col_header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        col_header.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
        row_header = self.calculs_table.verticalHeader()       
        row_header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        row_header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        row_header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        row_header.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
        row_header.setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        row_header.setSectionResizeMode(5, QtWidgets.QHeaderView.Stretch)

        self.calculs_layout.addWidget(self.calculs_table)
        self.browser.addWidget(self.calculs)

#####################################  div  ########################################################
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1092, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.timer = QtCore.QTimer(interval=1000, timeout=self.update_usages)
        self.timer.start()

        MainWindow.showMaximized()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "à baptiser"))
        self.cpu_label.setText(_translate("MainWindow", "CPU"))
        self.ram_label.setText(_translate("MainWindow", "RAM"))
        self.open.setText(_translate("MainWindow", "Ouvrir"))
        self.select_button.setText(_translate("MainWindow", "Sélection"))
        self.detection_button.setText(_translate("MainWindow", "Détection"))
        self.analyse_button.setText(_translate("MainWindow", "Analyse"))
        self.calculs_button.setText(_translate("MainWindow", "Calculs"))
        self.welcome_message_1.setText(_translate("MainWindow", "Bienvenue"))
        self.welcome_message_2.setText(_translate("MainWindow", "Veuillez ouvrir un dossier contenant vos images"))
        self.image_select_label.setText(_translate("MainWindow", "Confirmer les images à traiter"))
        self.type_1s_label.setText(_translate("MainWindow", "Echantillon 1 : Surnageant"))
        self.type_2s_label.setText(_translate("MainWindow", "Echantillon 2 : Surnageant"))
        self.type_2c_label.setText(_translate("MainWindow", "Echantillon 2 : Culot"))
        self.type_3s_label.setText(_translate("MainWindow", "Echantillon 3 : Surnageant"))
        self.type_1c_label.setText(_translate("MainWindow", "Echantillon 1 : Culot"))
        self.type_3c_label.setText(_translate("MainWindow", "Echantillon 3 : Culot"))
        self.cycle_next_button.setText(_translate("MainWindow", "Echantillon suivant"))
        self.cycle_prev_button.setText(_translate("MainWindow", "Echantillon précédent"))
        self.image_select_confirm.setText(_translate("MainWindow", "Confirmer"))
        self.show_conf_label.setText(_translate("MainWindow", "Montrer la confiance"))
        self.show_name_label.setText(_translate("MainWindow", "Montrer le nom du label"))
        self.taux_conf_label.setText(_translate("MainWindow", "Intervalle de confiance"))
        self.taux_conf.setSuffix(_translate("MainWindow", "%"))
        self.box_width_label.setText(_translate("MainWindow", "Épaisseur des boites"))
        self.start_detect_one.setText(_translate("MainWindow", "Détection sur cette image"))
        self.start_detect_many.setText(_translate("MainWindow", "Détection sur toutes les images"))
        self.show_detection_button.setText(_translate("MainWindow", "Voir les cellules détectées"))
        self.show_pred_button.setText(_translate("MainWindow", "Voir l'image"))
        self.start_calc_button.setText(_translate("MainWindow", "Lancer les calculs"))
        self.img_size_label.setText(_translate("MainWindow", "Grossissement des images :"))
        self.img_size_40.setText(_translate("MainWindow", "x40"))
        self.img_size_100.setText(_translate("MainWindow", "x100"))
        self.calc_button_save_xlsx.setText(_translate("MainWindow", "Sauvegarder au format .xlsx"))
        self.calc_button_save_csv.setText(_translate("MainWindow", "Sauvegarder au format .csv"))
        self.calculs_table.verticalHeaderItem(0).setText(_translate("MainWindow", "Nombre de cellules total"))
        self.calculs_table.verticalHeaderItem(1).setText(_translate("MainWindow", "Nombre capturées"))
        self.calculs_table.verticalHeaderItem(2).setText(_translate("MainWindow", "Nombre non-capturées"))
        self.calculs_table.verticalHeaderItem(3).setText(_translate("MainWindow", "Ratio de capture (%)"))
        self.calculs_table.verticalHeaderItem(4).setText(_translate("MainWindow", "Concentration culot (x10\u2075)"))
        self.calculs_table.verticalHeaderItem(5).setText(_translate("MainWindow", "Concentration surnageant (x10\u2075)"))
        self.calculs_table.horizontalHeaderItem(0).setText(_translate("MainWindow", "Echantillon 1"))
        self.calculs_table.horizontalHeaderItem(1).setText(_translate("MainWindow", "Echantillon 2"))
        self.calculs_table.horizontalHeaderItem(2).setText(_translate("MainWindow", "Echantillon 3"))
        self.calculs_table.horizontalHeaderItem(3).setText(_translate("MainWindow", "Écart Type"))
    
    def change_page_1(self):
        self.browser.setCurrentIndex(1)

    def change_page_2(self):
        self.browser.setCurrentIndex(2)
        if 'pred' in self.images[self.files.currentItem().text()]:
            self.set_image_from_cv(self.images[self.files.currentItem().text()]['image_pred'], 2)
        else:
            self.set_image_from_cv(self.images[self.files.currentItem().text()]['image'], 2)            

    def change_page_3(self):
        self.browser.setCurrentIndex(3)
        self.set_image_from_cv(self.images[self.files.currentItem().text()]['image_pred'], 3)

    def change_page_4(self):
        self.browser.setCurrentIndex(4)

    def openfile(self):
        self.files.clear()
        self.type_1s.clear()
        self.type_2s.clear()
        self.type_3s.clear()
        self.type_1c.clear()
        self.type_2c.clear()
        self.type_3c.clear()

        self.detection_button.setEnabled(False)
        self.analyse_button.setEnabled(False)
        self.calculs_button.setEnabled(False)
        self.cycle_next_button.setEnabled(False)
        self.cycle_prev_button.setEnabled(False)
        self.image_select_confirm.setEnabled(False)
        self.calc_button_save_csv.setEnabled(False)
        self.calc_button_save_xlsx.setEnabled(False)

        self.folder = str(QtWidgets.QFileDialog.getExistingDirectory(MainWindow, "Sélectionner un dossier"))
        if self.folder:
            for file in os.listdir(self.folder):
                if file.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.tif')):

                    name = re.findall('^.*(?=\.[^.]*$)', file.lower())[0]

                    if name[-2:] == '1s':
                        self.type_1s.addItem(file)
                    elif name[-2:] == '1c':
                        self.type_1c.addItem(file)
                    elif name[-2:] == '2s':
                        self.type_2s.addItem(file)
                    elif name[-2:] == '2c':
                        self.type_2c.addItem(file)
                    elif name[-2:] == '3s':
                        self.type_3s.addItem(file)
                    elif name[-2:] == '3c':
                        self.type_3c.addItem(file)

                    if name.find('100') != -1:
                        self.img_size_100.setChecked(True)
                    else:
                        self.img_size_40.setChecked(True)
        
        if self.type_1s.count() > 1:
            self.cycle_next_button.setEnabled(True)

        if self.type_1s.count() >= 1 and self.type_1c.count() >= 1 and self.type_2s.count() >= 1 and self.type_2c.count() >= 1 and self.type_3s.count() >= 1 and self.type_3c.count() >= 1:
            self.image_select_confirm.setEnabled(True)

        self.select_button.setEnabled(True)            
        self.change_page_1()

    def set_file_list(self):
        self.files.clear()
        self.files.addItem(self.type_1s.currentText())
        self.files.addItem(self.type_1c.currentText())
        self.files.addItem(self.type_2s.currentText())
        self.files.addItem(self.type_2c.currentText())
        self.files.addItem(self.type_3s.currentText())
        self.files.addItem(self.type_3c.currentText())

        self.loadimages()
        self.browser.setCurrentIndex(2)
        self.files.setCurrentRow(0)

    def cycle_next(self):
        if self.type_1s.count() > self.type_1s.currentIndex()+1 and self.type_1c.count() > self.type_1c.currentIndex()+1 and self.type_2s.count() > self.type_2s.currentIndex()+1 and self.type_2c.count() > self.type_2c.currentIndex()+1 and self.type_3s.count() > self.type_3s.currentIndex()+1 and self.type_3c.count() > self.type_3c.currentIndex()+1:
            self.type_1s.setCurrentIndex(self.type_1s.currentIndex() + 1)
            self.type_1c.setCurrentIndex(self.type_1c.currentIndex() + 1)
            self.type_2s.setCurrentIndex(self.type_2s.currentIndex() + 1)
            self.type_2c.setCurrentIndex(self.type_2c.currentIndex() + 1)
            self.type_3s.setCurrentIndex(self.type_3s.currentIndex() + 1)
            self.type_3c.setCurrentIndex(self.type_3c.currentIndex() + 1)
            self.cycle_prev_button.setEnabled(True)

        else:
            self.cycle_next_button.setEnabled(False)



    def cycle_prev(self):
        if self.type_1s.currentIndex() > 0:
            self.type_1s.setCurrentIndex(self.type_1s.currentIndex() - 1)
            self.type_1c.setCurrentIndex(self.type_1c.currentIndex() - 1)
            self.type_2s.setCurrentIndex(self.type_2s.currentIndex() - 1)
            self.type_2c.setCurrentIndex(self.type_2c.currentIndex() - 1)
            self.type_3s.setCurrentIndex(self.type_3s.currentIndex() - 1)
            self.type_3c.setCurrentIndex(self.type_3c.currentIndex() - 1)
            self.cycle_next_button.setEnabled(True)

        else: 
            self.cycle_prev_button.setEnabled(False)


    
    def loadimages(self):

        self.images = {}
        for item in [item.text() for item in [self.files.item(i) for i in range(self.files.count())]]:
            image = cv2.imdecode(np.fromfile(self.folder + '/' + item, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            self.images[item] = {'image' : image}
            self.images[item]['type'] = re.findall('^.*(?=\.[^.]*$)', item.lower())[0][-1]
            self.images[item]['echantillon'] = re.findall('^.*(?=\.[^.]*$)', item.lower())[0][-2]

        self.detection_button.setEnabled(True)
        self.analyse_button.setEnabled(False)
     
    def change_image(self):
        self.show_pred_button.setEnabled(False)
        self.show_detection_button.setEnabled(True)
        try:
            self.set_image_from_cv(self.images[self.files.currentItem().text()]['image_pred'])
            self.analyse_button.setEnabled(True)
        except:
            try:
                self.set_image_from_cv(self.images[self.files.currentItem().text()]['image'])
                self.browser.setCurrentIndex(2)
                self.analyse_button.setEnabled(False)
            except:
                pass
    

    def set_image_from_cv (self, img, page=0):
        height, width, channel = img.shape
        bytesPerLine = channel * width
        qImg = QtGui.QImage(img.data, width, height, bytesPerLine, QtGui.QImage.Format_RGB888)

        if page == 2:
            self.detect_viewer.setPhoto(QtGui.QPixmap(qImg))
        elif page == 3:
            self.analyse_viewer.setPhoto(QtGui.QPixmap(qImg))
        else:
            self.detect_viewer.setPhoto(QtGui.QPixmap(qImg))
            self.analyse_viewer.setPhoto(QtGui.QPixmap(qImg))

    def update_usages(self):
        self.cpu_bar.setValue(int(psutil.cpu_percent()))
        self.ram_bar.setValue(int(psutil.virtual_memory().percent))
    

    def get_pred(self, single = False):
        progress_dialog = QtWidgets.QProgressDialog('Détection en cours, patientez ...', None, 0, 0, MainWindow)
        progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
        progress_dialog.setWindowTitle('Détection')
        progress_dialog.show()

        if single:
            self.worker = DetectThread(self.images, self.taux_conf.value(), self.box_width.value(), self.show_conf.isChecked(), self.show_name.isChecked(), self.files.currentItem().text())
        else:
            self.worker = DetectThread(self.images, self.taux_conf.value(), self.box_width.value(), self.show_conf.isChecked(), self.show_name.isChecked())

        self.worker.finished.connect(self.handle_pred)
        self.worker.finished.connect(progress_dialog.close)
        self.worker.start()

    def handle_pred(self, result):
        self.images = result
        self.set_image_from_cv(self.images[self.files.currentItem().text()]['image_pred'], 2)
        self.analyse_button.setEnabled(True)

        all_done = True
        for key in self.images:
            if 'pred' not in self.images[key]:
                all_done = False

        if all_done:
            self.calculs_button.setEnabled(True)

    def show_detection(self):
        combined_image = utils.show_all_detections(self.images[self.files.currentItem().text()])
        self.set_image_from_cv(combined_image, 3)
        self.show_pred_button.setEnabled(True)
        self.show_detection_button.setEnabled(False)

    def show_pred(self):
        self.set_image_from_cv(self.images[self.files.currentItem().text()]['image_pred'], 3)
        self.show_pred_button.setEnabled(False)
        self.show_detection_button.setEnabled(True)

    def start_calc(self):
        self.images = utils.calcul_malassez(self.images, self.img_size_40.isChecked())
        self.set_calculs()

    def set_calculs(self):
        for key in self.images:

            if self.images[key]['type'] == 'c':
                self.calculs_table.setItem(1, int(self.images[key]['echantillon']) - 1, QtWidgets.QTableWidgetItem(str(len(self.images[key]['pred'].boxes.boxes))))
                self.calculs_table.setItem(4, int(self.images[key]['echantillon']) - 1, QtWidgets.QTableWidgetItem(str(self.images[key]['concentration'])))

            elif self.images[key]['type'] == 's':
                self.calculs_table.setItem(2, int(self.images[key]['echantillon']) - 1, QtWidgets.QTableWidgetItem(str(len(self.images[key]['pred'].boxes.boxes))))
                self.calculs_table.setItem(5, int(self.images[key]['echantillon']) - 1, QtWidgets.QTableWidgetItem(str(self.images[key]['concentration'])))

        e1_c_nb = int(self.calculs_table.item(1, 0).text())
        e2_c_nb = int(self.calculs_table.item(1, 1).text())
        e3_c_nb = int(self.calculs_table.item(1, 2).text())
        e1_s_nb = int(self.calculs_table.item(2, 0).text())
        e2_s_nb = int(self.calculs_table.item(2, 1).text())
        e3_s_nb = int(self.calculs_table.item(2, 2).text())

        e1_c_con = float(self.calculs_table.item(4, 0).text())
        e2_c_con = float(self.calculs_table.item(4, 1).text())
        e3_c_con = float(self.calculs_table.item(4, 2).text())
        e1_s_con = float(self.calculs_table.item(5, 0).text())
        e2_s_con = float(self.calculs_table.item(5, 1).text())
        e3_s_con = float(self.calculs_table.item(5, 2).text())

        e1_nb = e1_c_nb + e1_s_nb
        e2_nb = e2_c_nb + e2_s_nb
        e3_nb = e3_c_nb + e3_s_nb

        self.calculs_table.setItem(0, 0, QtWidgets.QTableWidgetItem(str(e1_nb)))
        self.calculs_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(e2_nb)))
        self.calculs_table.setItem(0, 2, QtWidgets.QTableWidgetItem(str(e3_nb)))

        e1_ratio = round(e1_c_nb / e1_nb * 100, 1)
        e2_ratio = round(e2_c_nb / e2_nb * 100, 1)
        e3_ratio = round(e3_c_nb / e3_nb * 100, 1)

        self.calculs_table.setItem(3, 0, QtWidgets.QTableWidgetItem(str(e1_ratio)))
        self.calculs_table.setItem(3, 1, QtWidgets.QTableWidgetItem(str(e2_ratio)))
        self.calculs_table.setItem(3, 2, QtWidgets.QTableWidgetItem(str(e3_ratio)))

        self.calculs_table.setItem(0, 3, QtWidgets.QTableWidgetItem(str(utils.get_std([e1_nb, e2_nb, e3_nb]))))
        self.calculs_table.setItem(1, 3, QtWidgets.QTableWidgetItem(str(utils.get_std([e1_c_nb, e2_c_nb, e3_c_nb]))))
        self.calculs_table.setItem(2, 3, QtWidgets.QTableWidgetItem(str(utils.get_std([e1_s_nb, e2_s_nb, e3_s_nb]))))
        self.calculs_table.setItem(3, 3, QtWidgets.QTableWidgetItem(str(utils.get_std([e1_ratio, e2_ratio, e3_ratio]))))
        self.calculs_table.setItem(4, 3, QtWidgets.QTableWidgetItem(str(utils.get_std([e1_c_con, e2_c_con, e3_c_con]))))
        self.calculs_table.setItem(5, 3, QtWidgets.QTableWidgetItem(str(utils.get_std([e1_s_con, e2_s_con, e3_s_con]))))

        self.calc_button_save_csv.setEnabled(True)
        self.calc_button_save_xlsx.setEnabled(True)

    def save_table_csv(self):
        import csv
        path, ok = QtWidgets.QFileDialog.getSaveFileName(MainWindow, 'Sauvegarder', os.getenv('HOME'), 'CSV(*.csv)')
        headers = [""] + [self.calculs_table.horizontalHeaderItem(i).text() for i in range(self.calculs_table.columnCount())]
        if ok:
            with open(path, "w", encoding = 'utf-8', newline="") as file:

                writer = csv.writer(file)

                writer.writerow(headers)

                for row in range(self.calculs_table.rowCount()):
                    row_data = [self.calculs_table.verticalHeaderItem(row).text()]
                    for column in range(self.calculs_table.columnCount()):
                        item = self.calculs_table.item(row, column)
                        row_data.append(item.text())

                    writer.writerow(row_data)

    def save_table_xlsx(self):
        from openpyxl import Workbook

        path, ok = QtWidgets.QFileDialog.getSaveFileName(MainWindow, 'Sauvegarder', os.getenv('HOME'), 'EXEL(*.xlsx)')

        if ok:
            workbook = Workbook()

            worksheet = workbook.active

            for row in range(self.calculs_table.rowCount()):
                for column in range(self.calculs_table.columnCount()):
                    item = self.calculs_table.item(row, column)
                    worksheet.cell(row=row+2, column=column+2, value=item.text())

            for column in range(self.calculs_table.columnCount()):
                header_cell = worksheet.cell(row=1, column=column+2)
                header_cell.value = self.calculs_table.horizontalHeaderItem(column).text()

            for row in range(self.calculs_table.rowCount()):
                header_cell = worksheet.cell(row=row+2, column=1)
                header_cell.value = self.calculs_table.verticalHeaderItem(row).text()

            workbook.save(filename=path)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
   
    QtCore.QDir.addSearchPath('Assets', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Assets'))
    app.setWindowIcon(QtGui.QIcon("Assets:Logo_small.png"))
    file = QtCore.QFile('Assets:Style.qss')
    file.open(QtCore.QFile.ReadOnly | QtCore.QFile.Text)
    app.setStyleSheet(str(file.readAll(), 'utf-8'))

    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())